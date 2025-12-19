from openpyxl import load_workbook
import pandas as pd

wb = load_workbook("lab5.xlsx", data_only=True)
ws = wb.active

data = []
column_letters = "ABCDFGHIJKLMNOPQRSTUVWX"

for row in ws.iter_rows(min_row=7, max_row=18):
    row_values = []
    for cell in row:
        if cell.column_letter not in column_letters:
            continue
        row_values.append(cell.value)
    data.append(row_values)

df = pd.DataFrame(data)
print(df)
